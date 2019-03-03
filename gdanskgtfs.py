import io
import os
import re
import csv
import json
import time
import zlib
import ftplib
import sqlite3
import zipfile
import argparse
import requests
import openpyxl
from bs4 import BeautifulSoup
from tempfile import NamedTemporaryFile
from datetime import date, datetime, timedelta

# Internal functions, to ease up parsing data

def _gettime(string):
    "Get GTFS-complaint time value from Gdańsk time value"
    if string.startswith("1899-12-30"):
        return(string.split("T")[-1])
    else:
        s = string.split("T")[-1].split(":")
        h = str(int(s[0]) + 24)
        return(":".join([h, s[1], s[2]]))

def _checkday(day):
    "Check if schedules are avilable for given date"
    timespans = json.loads(requests.get("http://91.244.248.30/dataset/c24aa637-3619-4dc2-a171-a23eec8f2172/resource/9c3d6fed-5394-4ef1-b2c6-c8716999149c/download/stoptimesspan.json").text)
    if "stopTimesSpans" in timespans: timespans["stopTimesSpans"]
    else: timespans["stopTimesSpan"]
    for agency in timespans:
        start = datetime.strptime(agency["startDate"], "%Y-%m-%d").date()
        end = datetime.strptime(agency["endDate"], "%Y-%m-%d").date()
        if start < end and (not (start <= day <= end)):
            return(False)
    return(True)

def _getrange(startdate):
    enum = 0
    while True:
        day = startdate + timedelta(days=enum)
        if not _checkday(day): break
        enum += 1
    return(range(enum))

def _gdyniaroutenames():
    zkm_website = requests.get("http://zkmgdynia.pl/")
    zkm_website.encoding = "utf-8"

    soup = BeautifulSoup(zkm_website.text, "html.parser")
    route_names = {}

    for route_div in soup.find_all("div", class_="nr_linii"):
        try: route = route_div.find("a", class_="nr_lini").text.strip()
        except AttributeError: continue

        try: name = soup.find("div", id="comment_"+route_div["id"]).text
        except (KeyError, AttributeError): continue

        name_pattern = list(map(str.strip, re.split(r"([<>-]{2,})", name)))

        for idx, name_part in enumerate(name_pattern):
            # Arrows
            if name_part == "<->":
                name_pattern[idx] = "—"

            elif name_part == "->":
                name_pattern[idx] = "→"

            elif name_part == "<-":
                name_pattern[idx] = "←"

            # Names
            else:

                # Get rid of brackets
                if "(" in name_part:
                    name_part = name_part.split("(")[0].strip()

                # Avoid something like "Gdynia: Gdynia Dworzec Gł. PKP"
                if ":" in name_part:
                    town_name, stop_name = map(str.strip, name_part.split(":"))

                    if set(town_name.split()).isdisjoint(stop_name.split()):
                        name_pattern[idx] = town_name + " " + stop_name

                    else:
                        name_pattern[idx] = stop_name

        route_names[route] = " ".join(name_pattern)

    return route_names

def _gdyniatownnames(ftp_login, ftp_pass):
    with open("ignore_temp_zkm_stops.xlsx", mode="w+b") as xlsx_file:
        with ftplib.FTP("ftp.zkmgdynia.pl", user=ftp_login, passwd=ftp_pass) as ftp:
            stop_files = sorted([i for i in ftp.nlst() if re.match(r"_wykaz_slupkow_\d{4}-\d{2}-\d{2}.xlsx", i)])
            for idx, file in enumerate(stop_files):
                if datetime.strptime(file, "_wykaz_slupkow_%Y-%m-%d.xlsx").date() > date.today():
                    if idx == 0: stop_file = stop_files[0]
                    else: stop_file = stop_files[idx - 1]
                    break
            else:
                stop_file = stop_files[-1]
            ftp.retrbinary("RETR "+stop_file, xlsx_file.write)

        xlsx_file.seek(0)
        stop_sheet = openpyxl.load_workbook("ignore_temp_zkm_stops.xlsx").active
        stop_town_names = {}

        for row in range(2, stop_sheet.max_row + 1):
            stop_name = stop_sheet.cell(row=row, column=3).value
            stop_town = stop_sheet.cell(row=row, column=6).value

            try: stop_id = int(stop_sheet.cell(row=row, column=2).value)
            except (ValueError, TypeError): stop_id = None

            if stop_town:
                if stop_name: stop_town_names[stop_name] = stop_town
                if stop_id: stop_town_names[str(stop_id)] = stop_town

    os.remove("ignore_temp_zkm_stops.xlsx")
    return stop_town_names

def _shouldaddtownname(stop_name, town_name):
    stop_name, town_name = map(str.upper, (stop_name, town_name))
    if town_name == "GDAŃSK": return False
    elif town_name in stop_name: return False
    for town_part_name in town_name.split(" "):
        if town_part_name in stop_name:
            return False
    return True

def _stopmergetable():
    merge_csv = requests.get("https://ckan.multimediagdansk.pl/dataset/c24aa637-3619-4dc2-a171-a23eec8f2172/resource/f8a5bedb-7925-40c9-8d66-dbbc830939b1/download/przystanki_wspolnegda_gdy.csv")
    merge_csv.encoding = "utf-8"

    table_buffer = io.StringIO(merge_csv.text)
    reader = csv.DictReader(table_buffer)

    merge_table = {}

    for row in reader:
        if row["main_organization_id"] == "1":
            merge_table[row["mapped_gmv_short_name"]] = row["main_gmv_short_name"]
        elif row["main_organization_id"] == "2":
            merge_table[row["main_gmv_short_name"]] = row["mapped_gmv_short_name"]

    table_buffer.close()
    return merge_table

# Parsing Scripts

def agencies(normalize):
    "Parse agencies to output/agency.txt GTFS file. If normalize is True only two will be created: ZTM Gdańsk and ZKM Gdynia."
    file = open("output/agency.txt", "w", encoding="utf-8", newline="\r\n")
    file.write("agency_id,agency_name,agency_url,agency_timezone,agency_lang\n")
    if normalize:
        file.write("99,ZTM Gdańsk,http://ztm.gda.pl,Europe/Warsaw,pl\n")
        file.write("98,ZKM Gdynia,http://zkmgdynia.pl,Europe/Warsaw,pl\n")
    else:
        agencies = json.loads(requests.get("http://91.244.248.30/dataset/c24aa637-3619-4dc2-a171-a23eec8f2172/resource/dff5f71f-0134-4ef3-8116-73c1a8e929a5/download/agency.json").text)
        agencies = agencies["agency"]
        for agency in agencies:
            agency_id = str(agency["agencyId"])
            agency_name = agency["agencyName"]
            agency_url = agency["agencyUrl"]
            agency_timezone = agency["agencyTimezone"]
            agency_lang = agency["agencyLang"]
            file.write(",".join([agency_id, agency_name, agency_url, agency_timezone, agency_lang + "\n"]))
    file.close()

def stops(startday, daysrange, ftp_login="", ftp_pass=""):
    "Parse stops for given day to output/stops.txt GTFS file"
    # Some variables
    allstops = json.loads(requests.get("http://91.244.248.30/dataset/c24aa637-3619-4dc2-a171-a23eec8f2172/resource/4c4025f0-01bf-41f7-a39f-d156d201b82b/download/stops.json").text)
    stopstable = {}
    stopattributes = {}

    stopRecalcAfterMaping = {}
    stopGdyGdaMaping = _stopmergetable()

    # Town names for ZKM Gdynia stops
    if ftp_login and ftp_pass: gdyniastoptowns = _gdyniatownnames(ftp_login, ftp_pass)
    else: gdyniastoptowns = {}

    # Database to merge stops over different days
    databaseconnection = sqlite3.connect(":memory:")
    databaseconnection.row_factory = sqlite3.Row
    database = databaseconnection.cursor()
    database.execute("CREATE TABLE stops (id text, orig_id text, name text, lat text, lon text, short_name text, merge_with text)")
    databaseconnection.commit()

    # Read stops to database
    for timediff in daysrange:
        day = startday + timedelta(days=timediff)
        try: stops = allstops[day.strftime("%Y-%m-%d")]["stops"]
        except KeyError: stops = allstops[day.strftime("%Y-%m-%d")]["node"]["stops"]
        for stop in stops:
            # Read data
            original_stop_id = str(stop["stopId"])
            stop_name = stop["stopDesc"].strip()
            stop_lat = str(stop["stopLat"])
            stop_lon = str(stop["stopLon"])

            # Check if stop should be merged:
            # stop_short_name isn't saved for GDY stops, because of merging that's done later
            if int(original_stop_id) < 30000: stop_short_name, merge_with = str(stop["stopShortName"]), None # GDA Stops
            else: stop_short_name, merge_with = None, stopGdyGdaMaping.get(str(stop["stopShortName"]), None) # GDY stops

            # Add town name for ZKM Gdynia stops
            if int(original_stop_id) >= 30000 and gdyniastoptowns:
                town_name = gdyniastoptowns.get(stop["stopShortName"]) or gdyniastoptowns.get(stop_name)
                if not town_name:
                    #pass
                    print("No town name for ZKM Gdynia stop {} ({})".format(original_stop_id, stop_name))
                elif  _shouldaddtownname(stop_name, town_name):
                    stop_name = town_name + " " + stop_name

            # Check against database
            database.execute("SELECT * FROM stops WHERE name=? AND lat=? AND lon=?", (stop_name, stop_lat, stop_lon))
            response = database.fetchone()

            if response:
                stop_id = response["id"]

            else:
                # Push into database
                response = database.execute("SELECT * FROM stops WHERE orig_id=?", (original_stop_id,))
                if response: used_ids = [response[x]["id"] for x in sorted(response)]
                else: used_ids = []
                stop_suffix = 0
                while True:
                    stop_id = "_".join([original_stop_id, str(stop_suffix)])
                    stop_suffix += 1
                    if stop_id not in used_ids: break
                database.execute("INSERT INTO stops VALUES (?,?,?,?,?,?,?)", (stop_id, original_stop_id, stop_name, stop_lat, stop_lon, stop_short_name, merge_with))

                # Save attributes
                stopattributes[stop_id] = {}
                stopattributes[stop_id]["virtual"] = stop["virtual"] == 1 or stop["nonpassenger"] == 1 or stop["depot"] == 1
                stopattributes[stop_id]["demand"] = stop["onDemand"] == 1

            databaseconnection.commit()
            if merge_with:
                if stop_id not in stopRecalcAfterMaping: stopRecalcAfterMaping[stop_id] = []
                stopRecalcAfterMaping[stop_id].append("-".join([day.strftime("%Y-%m-%d"), original_stop_id]))
            stopstable["-".join([day.strftime("%Y-%m-%d"), original_stop_id])] = stop_id

    # Export created database
    file = open("output/stops.txt", "w", encoding="utf-8", newline="\r\n")
    file.write("stop_id,stop_name,stop_lat,stop_lon\n")
    database.execute("SELECT * FROM stops")
    all_stops = database.fetchall()

    for stop in all_stops:
        stop_name = "\"" + stop["name"].replace("\"", "\"\"").replace("\'", "\"\"") + "\""

        # Check if stop could be merged
        if stop["merge_with"]:
            database.execute("SELECT * FROM stops WHERE short_name=?", (stop["merge_with"], ))

            # Now check if it can be merged
            merge_with = database.fetchone()

            # If it can be: merge and rewrite value in stopstable
            if merge_with:
                #print("Merging common stop: GDY {} -> GDA {} ({})".format(stop["id"], merge_with["id"], merge_with["name"]))
                for recalculate_mapped in stopRecalcAfterMaping[stop["id"]]:
                    stopstable[recalculate_mapped] = merge_with["id"]

            # If not just print it to stops.txt
            else:
                file.write(",".join([stop["id"], stop_name, stop["lat"], stop["lon"]]) + "\n")

        # If not, just write to stops.txt
        else:
            file.write(",".join([stop["id"], stop_name, stop["lat"], stop["lon"]]) + "\n")

    file.close()
    databaseconnection.commit()

    return stopstable, stopattributes

def routes(startday, daysrange, normalize):
    "Parse routes for given day to output/routes.txt GTFS file. If normalize is True, then agency_id will be filtered to ZTM or ZKM."
    # Some variables
    allroutes = json.loads(requests.get("http://91.244.248.30/dataset/c24aa637-3619-4dc2-a171-a23eec8f2172/resource/22313c56-5acf-41c7-a5fd-dc5dc72b3851/download/routes.json").text)
    routeslist = {}
    routestable = {}

    # Database to merge same routes over different days
    databaseconnection = sqlite3.connect(":memory:")
    databaseconnection.row_factory = sqlite3.Row
    database = databaseconnection.cursor()
    database.execute("CREATE TABLE routes (id text, orig_id text, agency text, short_name text, long_name text, type text, color text)")
    databaseconnection.commit()

    # Read routes to database
    for timediff in daysrange:
        day = startday + timedelta(days=timediff)
        routeslist[day.strftime("%Y-%m-%d")] = []
        gdyniaroutes = _gdyniaroutenames()
        try: routes = allroutes[day.strftime("%Y-%m-%d")]["routes"]
        except KeyError: routes = allroutes[day.strftime("%Y-%m-%d")]["node"]["routes"]
        for route in routes:
            # Read data
            agency_id = str(route["agencyId"])
            original_route_id = str(route["routeId"])
            route_short_name = str(route["routeShortName"])
            route_long_name = route["routeLongName"] if route["routeShortName"] != route["routeLongName"] else ""

            route_long_name = route_long_name.replace(" - ", " — ")

            if agency_id == "2":
                #Gdańsk Trams
                route_type = "0"
                route_color = "BB0000,FFFFFF"
            elif agency_id == "5":
                #Gdynia Trolleybuses
                route_type = "800"
                route_color = "11CC11,000000"
            elif route_short_name.startswith("N"):
                #Night Buses
                route_type = "3"
                route_color = "000000,FFFFFF"
            elif not route_short_name.isnumeric():
                #Express Busses
                route_type = "3"
                route_color = "FFCC22,000000"
            else:
                #Normal Busses
                route_type = "3"
                route_color = "2222BB,FFFFFF"
            if normalize:
                if 10000 <= int(original_route_id) < 11000: agency_id = "98"
                else: agency_id = "99"

            # Gdynia route_long_name
            if 10000 <= int(original_route_id) < 11000 and route_short_name in gdyniaroutes:
                route_long_name = gdyniaroutes[route_short_name]

            # Check against database
            database.execute("SELECT * FROM routes WHERE agency=? AND short_name=? AND long_name=?", (agency_id, route_short_name, route_long_name))
            response = database.fetchone()
            if response:
                route_id = response["id"]
            else:
                response = database.execute("SELECT * FROM routes WHERE orig_id=?", (original_route_id,))
                if response: used_ids = [response[x]["id"] for x in sorted(response)]
                else: used_ids = []
                route_suffix = 0
                while True:
                    route_id = "_".join([original_route_id, str(route_suffix)])
                    route_suffix += 1
                    if route_id not in used_ids: break
                database.execute("INSERT INTO routes VALUES (?,?,?,?,?,?,?)", (route_id, original_route_id, agency_id, route_short_name, route_long_name, route_type, route_color))
            databaseconnection.commit()
            routeslist[day.strftime("%Y-%m-%d")].append(original_route_id)
            routestable["-".join([day.strftime("%Y-%m-%d"), original_route_id])] = route_id

    # Export created databse
    file = open("output/routes.txt", "w", encoding="utf-8", newline="\r\n")
    file.write("agency_id,route_id,route_short_name,route_long_name,route_type,route_color,route_text_color\n")
    database.execute("SELECT * FROM routes")
    for route in database.fetchall():
        route_name = "\"" + route["long_name"].replace("\r", "").replace("\n", "").replace("\"", "\"\"").replace("\'", "\"\"") + "\""
        file.write(",".join([route["agency"], route["id"], route["short_name"], route_name, route["type"], route["color"] + "\n"]))
    file.close()
    databaseconnection.commit()
    return routeslist, routestable

def times(startday, daysrange, routeslist, routestable, stopstable, stopattributes):
    "Parse stop_times for given day to output/stop_times.txt and output/trips.txt GTFS file"
    fileTimes = open("output/stop_times.txt", "w", encoding="utf-8", newline="\r\n")
    fileTimes.write("trip_id,arrival_time,departure_time,stop_id,original_stop_id,stop_sequence,pickup_type,drop_off_type\n")
    fileTrips = open("output/trips.txt", "w", encoding="utf-8", newline="\r\n")
    fileTrips.write("service_id,route_id,original_route_id,trip_id,wheelchair_accessible\n")
    for timediff in daysrange:
        day = (startday + timedelta(days=timediff)).strftime("%Y-%m-%d")
        for route in routeslist[day]:
            print("\033[1A\033[KParsing stop_times: Day %s, route %s" % (day, route))
            trips = {}
            try:
                times = json.loads(requests.get("http://87.98.237.99:88/stopTimes?date=%s&routeId=%s" % (day, route)).text)
                times = times["stopTimes"]
            except (json.decoder.JSONDecodeError, KeyError):
                continue

            # Load times
            for time in times:
                trip_id = "R%sD%sT%sS%sO%s" % (str(time["routeId"]), day, time["tripId"], time["busServiceName"], str(time["order"]))
                stop_id = stopstable[day + "-" + str(time["stopId"])]
                arrival_time = _gettime(time["arrivalTime"])
                departure_time = _gettime(time["departureTime"])

                if trip_id not in trips.keys():
                    trips[trip_id] = {"data": {"route": str(time["routeId"]), "low_floor": ""}, "times": []}

                # Pick/Drop type
                if time["virtual"] == 1 or time["nonpassenger"] == 1 or stopattributes[stop_id]["virtual"]:
                    # Stops not-for-passengers won't be included
                    continue
                elif time["onDemand"] == 1 or stopattributes[stop_id]["demand"]:
                    pd_type = "3,3"
                else:
                    pd_type = "0,0"

                # Wheelchair Accessibility
                if not trips[trip_id]["data"]["low_floor"]:
                    if time["wheelchairAccessible"] == 1:
                        trips[trip_id]["data"]["low_floor"] = "1"
                    elif time["wheelchairAccessible"] == 0:
                        trips[trip_id]["data"]["low_floor"] = "2"
                    else:
                        trips[trip_id]["data"]["low_floor"] = "0"

                # Append to trips
                trips[trip_id]["times"].append({ \
                    "arrival": arrival_time, "departure": departure_time,
                    "stop": str(time["stopId"]), "stop_seq": time["stopSequence"],
                    "pd_type": pd_type
                })

            # Export times
            for trip_id, trip_info in trips.items():

                # Don't export one-stop and two-stop trips [as ZTM suggests upon email comunication]
                if len(trip_info["times"]) <= 2:
                    continue

                # Sort times by stop_sequence
                trip_info["times"] = sorted(trip_info["times"], key=lambda i: i["stop_seq"])

                # Dump trip data
                fileTrips.write(",".join([
                    day, #service_id
                    routestable[day + "-" + trip_info["data"]["route"]], #route_id
                    trip_info["data"]["route"], #original_route_id
                    trip_id,
                    trip_info["data"]["low_floor"] # wheelchair_accessible
                ]) + "\n")

                # Dump times data
                for time in trip_info["times"]:
                    fileTimes.write(",".join([
                        trip_id, time["arrival"], time["departure"], stopstable[day + "-" + time["stop"]],
                        time["stop"], str(time["stop_seq"]), time["pd_type"]
                    ]) + "\n")




    print("\033[1A\033[KParsing stop_times")
    fileTrips.close()
    fileTimes.close()

def calendar(startday, daysrange, extenddates):
    "Create calendar_dates file for provided days"
    file = open("output/calendar_dates.txt", "w", encoding="utf-8", newline="\r\n")
    file.write("date,service_id,exception_type\n")
    services = {}
    for timediff in daysrange:
        day = (startday + timedelta(days=timediff))
        daystr = day.strftime("%Y-%m-%d")
        services[day.weekday()] = day
        file.write("%s,%s,1\n" % (daystr.replace("-", ""), daystr))
    if extenddates:
        for timediff in range(max(daysrange)+1, 31):
            day = (startday + timedelta(days=timediff))
            service = services[day.weekday()].strftime("%Y-%m-%d")
            file.write("%s,%s,1\n" % (day.strftime("%Y%m%d"), service))
    file.close()

def feedinfo(startday, daysrange, extenddates):
    "Create feed_info in output/feed_info.txt to fulfil licencing needs"
    endday = (startday + timedelta(days=max(daysrange))).strftime("%Y%m%d") if not extenddates else (startday + timedelta(days=30)).strftime("%Y%m%d")
    startday = startday.strftime("%Y%m%d")
    file = open("output/feed_info.txt", "w", encoding="utf-8", newline="\r\n")
    file.write("feed_publisher_name,feed_publisher_url,feed_lang,feed_start_date,feed_end_date,feed_version\n")
    file.write("Zarząd Transportu Miejskiego w Gdańsku,\"http://91.244.248.30/dataset/tristar\",pl,%s,%s,%s\n" % (startday, endday, date.today().strftime("%Y-%m-%d")))
    file.close()

# Utility Scripts

def cleanup():
    "Cleans output/ directory before parsing."
    if not os.path.exists("output"): os.mkdir("output")
    for file in [os.path.join("output", x) for x in os.listdir("output")]: os.remove(file)

def tables(routestable, stopstable):
    "Exports routes and stops tables in order to match static with RT data to tables.json"
    file = open("tables.json", "w", encoding="utf-8", newline="\r\n")
    file.write(json.dumps({"routes": routestable, "stops": stopstable}, sort_keys=True, indent=4))
    file.close()

def zip():
    "Zips the content of output/*.txt to gtfs.zip"
    with zipfile.ZipFile("gtfs.zip", mode="w", compression=zipfile.ZIP_DEFLATED) as archive:
        for file in os.listdir("output"):
            if file.endswith(".txt"):
                archive.write("output/" + file, arcname=file)

# Main Funcionlity

def gdanskgtfs(day=date.today(), normalize=False, exporttables=False, extenddates=False, zkm_ftp_login="", zkm_ftp_pass=""):
    daysrange = _getrange(day)
    #daysrange = range(1)
    if daysrange:
        print("Downloading schedules for %s to %s" % (day.strftime("%Y-%m-%d"), (day + timedelta(max(daysrange))).strftime("%Y-%m-%d")))
        print("Cleaning up output/ dir")
        cleanup()

        print("Parsing agencies")
        agencies(normalize)

        print("Creating calendar and feed_info")
        calendar(day, daysrange, extenddates)
        feedinfo(day, daysrange, extenddates)

        print("Parsing stops")
        stable, sattrib  = stops(day, daysrange, zkm_ftp_login, zkm_ftp_pass)

        print("Parsing routes")
        rlist, rtable = routes(day, daysrange, normalize)

        print("Parsing stop_times")
        times(day, daysrange, rlist, rtable, stable, sattrib)

        print("Zipping to gtfs.zip")
        zip()

        if exporttables:
            print("Exporting routes and stops tables to tables.json")
            tables(rtable, stable)

    else:
        print("Error! Full schedules are not available for date %s!" % day.strftime("%Y-%m-%d"))

if __name__ == "__main__":
    st = time.time()
    argprs = argparse.ArgumentParser()
    argprs.add_argument("-e", "--extend", action="store_true", required=False, dest="extend", help="artifically extend effective dates to 30 days")
    argprs.add_argument("-t", "--tables", action="store_true", required=False, dest="tables", help="export routes and stops tables to tables.json")
    argprs.add_argument("-n", "--normalize", action="store_true", required=False, dest="normalize", help="normalize agencies to ZTM Gdańsk and ZKM Gdynia")
    argprs.add_argument("-d", "--day", default="", required=False, metavar="YYYY-MM-DD", dest="day", help="the start day for which the feed should start")
    argprs.add_argument("-zl", "--zkm-ftp-login", default="", required=False, dest="zkm_ftp_login", help="login for ZKM Gdynia FTP server (ftp://ftp.zkmgdynia.pl/), to get fixed stop names")
    argprs.add_argument("-zp", "--zkm-ftp-pass", default="", required=False, dest="zkm_ftp_pass", help="password for ZKM Gdynia FTP server")

    args = vars(argprs.parse_args())
    if args["day"]: day = datetime.strptime(args["day"], "%Y-%m-%d").date()
    else: day = date.today()
    print("""
  __                    __ ___ _  __
 /__  _|  _. ._   _ |  /__  | |_ (_
 \_| (_| (_| | | _> |< \_|  | |  __)
    """)
    gdanskgtfs(day, args["normalize"], args["tables"], args["extend"], args["zkm_ftp_login"], args["zkm_ftp_pass"])
    print("=== Done! In %s sec. ===" % round(time.time() - st, 3))
